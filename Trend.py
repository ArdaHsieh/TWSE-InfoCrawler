# -*- coding: utf-8 -*-
"""
Created on Tue Nov 28 2017

@author: I-Ta Hsieh(Arda)
"""

import requests
import json
from bs4 import BeautifulSoup
import os
import time
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, colors


TWSE_Price = 0        # 台股大盤指數
TWSE_UD = 0           # 台股大盤漲跌
TWSE_UDR = 0          # 台股大盤漲跌幅
TWSE_Vol = 0          # 台股大盤成交量
TWSE_FBS = 0          # 台股大盤外資買賣超
INDU_Price = 0        # 道瓊工業指數
INDU_UDR = 0          # 道瓊工業指數漲跌幅
NAS_Price = 0         # Nasdaq指數
NAS_UDR = 0           # Nasdaq指數漲跌幅
SP500_Price = 0       # S&P 500指數
SP500_UDR = 0         # S&P 500指數漲跌幅
SOX_Price = 0         # 費城半導體指數
SOX_UDR = 0           # 費城半導體指數漲跌幅
USDEx = 0             # 美金兌台幣匯率
USDEx_UD = 0          # 美金兌台幣匯率漲跌
FU = 0                # 台股期貨外資留倉口數
DueFu = 0             # 期貨結算日外資留倉口數
TM_5_Bull = 0         # 台股期貨前5大交易人近月留倉部位
TM_10_Bull = 0        # 台股期貨前10大交易人近月留倉部位
ALL_5_Bull = 0        # 台股期貨前5大交易人所有留倉部位
ALL_10_Bull = 0       # 台股期貨前10大交易人所有留倉部位
RIBS_Ratio = 0        # 散戶多空比
Buy_Call = 0          # 外資選擇權未平倉(買權)
Buy_Put = 0           # 外資選擇權未平倉(賣權)
PC_Ratio = 0          # P/C Ratio
Tx_Due_yyyy = ""      # 期貨結算(年)
Tx_Due_mm = ""        # 期貨結算(月)
Tx_Due_dd = ""        # 期貨結算(日)


def get_url(url):
    html = requests.get(url).text.encode('utf-8-sig')
    
    return html

def post_url(url, payloads):
    html = requests.post(url, data = payloads).text.encode('utf-8-sig')
    
    return html


def string_to_nums(text):
    number = ""
    numbers = ["1","2","3","4","5","6","7","8","9","0",".","-"]
    for char in  text:
        if char in numbers:
            number += char
        elif char == "\\" or (char == " "  and number != "") or char == "(": 
            break
    
    return float(number)


# 台股大盤(TSE) 
def TWSE(yyyy, mm, dd):
    global TWSE_Price, TWSE_UD, TWSE_UDR, TWSE_Vol, TWSE_FBS
    
    # 大盤指數
    url_TWSE = "http://www.twse.com.tw/exchangeReport/MI_INDEX?response=json&date=" + yyyy + mm + dd
    html_TWSE = get_url(url_TWSE)
    TWSE_data1 = json.loads(html_TWSE)
    TWSE_data2 = TWSE_data1['data1']
    
    TWSE_Price = string_to_nums(TWSE_data2[1][1])    # 加權指數
    TWSE_UD = string_to_nums(TWSE_data2[1][3])       # 漲跌
    try:                                             # 漲跌幅
        TWSE_UDR = string_to_nums(TWSE_data2[1][4])      
    except:
        TWSE_UDR = 0.0
    if TWSE_UDR < 0:
        TWSE_UD = -TWSE_UD
    else:
        TWSE_UD = TWSE_UD
    
    # 大盤成交量
    Voldate = str(int(yyyy)-1911) + "/" + mm + "/" + dd
    url_TWSE_Vol = "http://www.twse.com.tw/exchangeReport/FMTQIK?response=json&date=" + yyyy + mm + dd
    html_TWSE_Vol = get_url(url_TWSE_Vol)
    TWSE_Vol_data1 = json.loads(html_TWSE_Vol)
    TWSE_Vol_data2 = TWSE_Vol_data1['data']
    for Voldata in TWSE_Vol_data2:
        if Voldata[0] == Voldate:
            TWSE_Vol = string_to_nums(Voldata[2])/100000000    # 成交量
            break
    TWSE_Vol = "%.2f" % TWSE_Vol
    
    # 外資買賣超
    url_TWSE_Foreign = "http://www.twse.com.tw/fund/BFI82U?response=json&dayDate=" + yyyy + mm + dd
    html_Foreign = get_url(url_TWSE_Foreign)
    TWSE_Foreign_data1 = json.loads(html_Foreign)
    TWSE_Foreign_data2 = TWSE_Foreign_data1['data']
    
    TWSE_FBS = string_to_nums(TWSE_Foreign_data2[3][3])/100000000
    TWSE_FBS = "%.2f" % TWSE_FBS
    

# 美股
def USASEx(yyyy, mm, dd):
    global INDU_Price, INDU_UDR, NAS_Price, NAS_UDR, SP500_Price, SP500_UDR, SOX_Price, SOX_UDR
    
    f = open('USEQ_Date.txt', 'r')
    dates = [date.rstrip() for date in f.readlines()]
    
    if (yyyy + '/' + mm + '/' + dd) in dates:
        INDU_Price = '休市'
        INDU_UDR = 0
        NAS_Price = '休市'
        NAS_UDR = 0
        SP500_Price = '休市'
        SP500_UDR = 0
        SOX_Price = '休市'
        SOX_UDR = 0
    else:
        url = "http://www.stockq.org/stock/history/" + yyyy +"/" + mm + "/" + yyyy + mm + dd + "_tc.php"
        html = get_url(url)
    
        soup = BeautifulSoup(html, 'html.parser')
        table = soup.find_all('table', {'class':'marketdatatable'})
        tr = table[2].find_all('tr')
    
        # 道瓊工業指數
        INDU = tr[2].find_all('td')
        INDU_Price = string_to_nums(INDU[2].text)    # 指數
        INDU_UDR = string_to_nums(INDU[3].text)      # 漲跌幅
 
        # NASDAQ指數
        NAS = tr[10].find_all('td')
        NAS_Price = string_to_nums(NAS[2].text)    # 指數
        NAS_UDR = string_to_nums(NAS[3].text)      # 漲跌幅
    
        # S&P500指數 
        SP500 = tr[4].find_all('td')
        SP500_Price = string_to_nums(SP500[2].text)    # 指數
        SP500_UDR = string_to_nums(SP500[3].text)      # 漲跌幅
    
        # 費城半導體指數
        SOX = tr[11].find_all('td')
        SOX_Price = string_to_nums(SOX[2].text)    # 指數
        SOX_UDR = string_to_nums(SOX[3].text)      # 漲跌幅

    
# 美金匯率
def USD_NTD(yyyy, mm, dd):
    global USDEx, USDEx_UD
    
    url = "http://www.stockq.org/stock/history/" + yyyy +"/" + mm + "/" + yyyy + mm + dd + "_tc.php"
    html = get_url(url)
    
    soup = BeautifulSoup(html, 'html.parser')
    table = soup.find_all('table', {'class':'marketdatatable'})
    tr = table[6].find_all('tr')

    USDEx = string_to_nums(tr[17].find_all('td')[1].text)
    USDEx_UD = string_to_nums(tr[17].find_all('td')[2].text)
    

# 期貨未平倉
def TWFU(yyyy, mm, dd):
    url = "https://www.taifex.com.tw/cht/3/futContractsDate"
    payloads = {"queryType":1, "doQuery":1, "queryDate":yyyy+ '/' + mm + '/' + dd}
    html = post_url(url, payloads)
    
    soup = BeautifulSoup(html, 'html.parser')
    tr = soup.find_all('tr', {'class':'12bk'})
    
    # 台指期
    TX_data = tr[5].find_all('td')
    TX = string_to_nums(TX_data[11].text)
    
    # 小台指
    MTX_data = tr[14].find_all('td')
    MTX = string_to_nums(MTX_data[11].text)
    
    FU = TX + MTX/4.0
    
    return FU


def TWFUOC(yyyy, mm, dd):
    global FU, DueFu, Tx_Due_yyyy, Tx_Due_mm, Tx_Due_dd
    
    FU = int(TWFU(yyyy, mm, dd))
    if Tx_Due_yyyy == yyyy and Tx_Due_mm == mm and Tx_Due_dd == dd:
        DueFu = "結算日"
    else:
        DueFu = FU - int(TWFU(Tx_Due_yyyy, Tx_Due_mm, Tx_Due_dd)) 
    

# 期貨5/10大交易人留倉部位
def TWF_5_10_UOC(yyyy, mm, dd):
    global TM_5_Bull, TM_10_Bull, ALL_5_Bull, ALL_10_Bull
    
    url = "https://www.taifex.com.tw/cht/3/largeTraderFutQry"
    payloads = {"queryDate":yyyy+ '/' + mm + '/' + dd, "contractId":"TX"}
    html = post_url(url, payloads)
    
    soup = BeautifulSoup(html, 'html.parser')
    table = soup.find_all('table')
    table_f = table[2].find_all('table', {'class':'table_f'})
    tr = table_f[0].find_all('tr')
    
    TM_data = tr[4].find_all('td', {'class':'11b'})
    ALL_data = tr[5].find_all('td', {'class':'11b'})
    
    TM_5_Bull = int(string_to_nums(TM_data[1].text.split(' ')[1]) - string_to_nums(TM_data[5].text.split(' ')[1]))
    TM_10_Bull = int(string_to_nums(TM_data[3].text.split(' ')[1]) - string_to_nums(TM_data[7].text.split(' ')[1]))
    ALL_5_Bull = int(string_to_nums(ALL_data[1].text.split(' ')[1]) - string_to_nums(ALL_data[5].text.split(' ')[1]))
    ALL_10_Bull = int(string_to_nums(ALL_data[3].text.split(' ')[1]) - string_to_nums(ALL_data[7].text.split(' ')[1]))
    
    
# 小台散戶多空比
def TWMTX(yyyy, mm, dd):
    global RIBS_Ratio
    # 小台全部留倉
    url = "https://www.taifex.com.tw/cht/3/futDailyMarketReport"
    payloads = {"queryType":2, "marketCode":0, "commodity_id":"MTX", "queryDate":yyyy+ '/' + mm + '/' + dd,"MarketCode": 0,"commodity_idt": "MTX"}
    html_RI = post_url(url, payloads)
    
    soup_I = BeautifulSoup(html_RI, 'html.parser')
    table_f_I = soup_I.find_all('table', {'class': 'table_f'})
    tr_I = table_f_I[0].find_all('tr')
    td_I = tr_I[-1].find_all('td' , {'class': '12bk'})
    MTXO = string_to_nums(td_I[4].text)
    
    # 三大法人留倉
    url_II = "https://www.taifex.com.tw/cht/3/futContractsDate"
    payloads_II = {"queryType":1,"doQuery":1,"queryDate":yyyy+ '/' + mm + '/' + dd,"commodityId":"MXF"}
    html_II = post_url(url_II, payloads_II)
    
    soup_II = BeautifulSoup(html_II, 'html.parser')
    table_f_II = soup_II.find_all('table', {'class': 'table_f'})
    tr_II = table_f_II[0].find_all('tr', {'class': '12bk'})
    
    SI_data = tr_II[3].find_all('td')    # 自營商
    IT_data = tr_II[4].find_all('td')    # 投信
    FI_data = tr_II[5].find_all('td')    # 外資
    # 空 - 多
    SI_OC = string_to_nums(SI_data[11].text) - string_to_nums(SI_data[9].text)  
    IT_OC = string_to_nums(IT_data[9].text) - string_to_nums(IT_data[7].text)
    FI_OC = string_to_nums(FI_data[9].text) - string_to_nums(FI_data[7].text)
    MTX_II_OC = SI_OC + IT_OC + FI_OC
    
    RIBS_Ratio = (MTX_II_OC/MTXO)*100  
    RIBS_Ratio = "%.2f" % RIBS_Ratio
    

# 選擇權
def TWOP(yyyy, mm, dd):
    global Buy_Call, Buy_Put
    
    url = "https://www.taifex.com.tw/cht/3/optContractsDate"
    payloads = {"queryType":1,"doQuery":1,"queryDate":yyyy+ '/' + mm + '/' + dd,"commodityId":"TXO"}
    html = post_url(url, payloads)
    
    soup = BeautifulSoup(html, 'html.parser')
    table_f = soup.find_all('table', {'class': 'table_f'})
    tr = table_f[0].find_all('tr', {'class': '12bk'})
    td = tr[5].find_all('td')
    
    # 買進買權
    #Buy_Call_data = data3[5].find_all('td')
    Buy_Call = string_to_nums(td[8].text)/100000
    Buy_Call = "%.2f" % Buy_Call
    
    # 買進賣權
    #Buy_Put_data = data3[8].find_all('td')
    Buy_Put = string_to_nums(td[10].text)/100000
    Buy_Put = "%.2f" % Buy_Put
    

# 選擇權 P/C Ratio
def TWPCR(yyyy, mm, dd):
    global PC_Ratio 
    url = "https://www.taifex.com.tw/cht/3/pcRatio"
    payloads = {"queryStartDate":yyyy + "/" + mm + "/" + dd, "queryEndDate":yyyy + "/" + mm + "/" + dd}
    html = post_url(url, payloads)
    
    soup = BeautifulSoup(html, 'html.parser')
    table_a = soup.find_all('table', {'class':'table_a'})
    tr = table_a[0].find_all('tr')
    td = tr[1].find_all('td')
    
    # P/C Ratio
    PC_Ratio = string_to_nums(td[-1].text)


# 期貨結算日資訊
def TxDueData(yyyy, mm, dd):
    global Tx_Due_yyyy, Tx_Due_mm, Tx_Due_dd
    
    with open ("TX Due Date.txt", 'r', encoding = 'UTF-8-sig') as f:
        Date = f.read()
        Due_Date = Date.split()
        
        for i in range(len(Due_Date)):
            if int(yyyy) > int(Due_Date[i][0:4]):
                Tx_Due_Date = Due_Date[i]
            elif int(yyyy) == int(Due_Date[i][0:4]) and int(mm) > int(Due_Date[i][5:7]):
                Tx_Due_Date = Due_Date[i]
            elif int(yyyy) == int(Due_Date[i][0:4]) and int(mm) == int(Due_Date[i][5:7]) and int(dd) >= int(Due_Date[i][8:10]):
                Tx_Due_Date = Due_Date[i]
        
        Tx_Due_yyyy = Tx_Due_Date[0:4]
        Tx_Due_mm = Tx_Due_Date[5:7]
        Tx_Due_dd = Tx_Due_Date[8:10]
        
        
# 判斷當日有無交易
def Y_N_Market(yyyy, mm, dd):
    url_TWSE = "http://www.twse.com.tw/exchangeReport/MI_INDEX?response=json&date=" + yyyy + mm + dd
    html_TWSE = get_url(url_TWSE)
    TWSE_data1 = json.loads(html_TWSE)
    
    return len(TWSE_data1), TWSE_data1["stat"]


# Display
def Disp(yyyy, mm, dd):
    print("----------------------------------------------------")
    print("日期: {}/{}/{}".format(yyyy, mm, dd))
    print("台股加權指數: {}".format(TWSE_Price))
    print("台股加權指數漲跌: {}".format(TWSE_UD))
    print("台股加權指數漲跌幅: {}%".format(TWSE_UDR))
    print("台股加權指數成交量: {}億".format(TWSE_Vol))
    print("台股加權指數外資買賣超: {}億".format(TWSE_FBS))
    print("道瓊工業指數: {}".format(INDU_Price))
    print("道瓊工業指數漲跌幅: {}%".format(INDU_UDR))
    print("Nasdaq指數: {}".format(NAS_Price))
    print("Nasdaq指數漲跌幅: {}%".format(NAS_UDR))
    print("S&P 500指數: {}".format(SP500_Price))
    print("S&P 500指數漲跌幅: {}%".format(SP500_UDR))
    print("費城半導體指數: {}".format(SOX_Price))
    print("費城半導體指數漲跌幅: {}%".format(SOX_UDR))
    print("新台幣匯率: {}".format(USDEx))
    print("新台幣匯率漲跌: {}".format(USDEx_UD))
    print("外資期貨未平倉口數: {}".format(FU))
    print("未平平倉口數與結算日相比: {}".format(DueFu))
    print("前5大交易人留倉部位(所有): {}".format(ALL_5_Bull))
    print("前10大交易人留倉部位(所有): {}".format(ALL_10_Bull))
    print("前5大交易人留倉部位(當月): {}".format(TM_5_Bull))
    print("前10大交易人留倉部位(當月): {}".format(TM_10_Bull))
    print("前5大交易人未來指標(所有-當月): {}".format(str(int(ALL_5_Bull)-int(TM_5_Bull))))
    print("前10大交易人未來指標(所有-當月): {}".format(str(int(ALL_10_Bull)-int(TM_10_Bull))))
    print("外資選擇權未平倉金額(買權/賣權): {}億/{}億".format(Buy_Call, Buy_Put))
    print("選擇權 P/C Ratio 未平倉量: {}".format(PC_Ratio))
    print("散戶多空比: {}%".format(RIBS_Ratio))
    print("----------------------------------------------------\n")

# Excel漲跌上色
def Excel_color(Row_num):
    file_name = "Stock Analysis.xlsx"
    ftRed = Font(color = colors.RED)
    ftGreen = Font(color = colors.GREEN)
    
    wb = load_workbook(file_name)
    ws = wb.active
    
    Info_TWSE_Price = ws.cell(row=Row_num, column=2)
    Info_TWSE_UD = ws.cell(row=Row_num, column=3)
    Info_TWSE_UDR = ws.cell(row=Row_num, column=4)
    Info_INDU_Price = ws.cell(row=Row_num, column=7)
    Info_INDU_UDR = ws.cell(row=Row_num, column=8)
    Info_NAS_Price = ws.cell(row=Row_num, column=9)
    Info_NAS_UDR = ws.cell(row=Row_num, column=10)
    Info_SP500_Price = ws.cell(row=Row_num, column=11)
    Info_SP500_UDR = ws.cell(row=Row_num, column=12)
    Info_SOX_Price = ws.cell(row=Row_num, column=13)
    Info_SOX_UDR = ws.cell(row=Row_num, column=14)
    
    # 台股上色
    if float(Info_TWSE_UD.value) > 0:
        Info_TWSE_Price.font = ftRed
        Info_TWSE_UD.font = ftRed
        Info_TWSE_UDR.font = ftRed
    elif float(Info_TWSE_UD.value) < 0:
        Info_TWSE_Price.font = ftGreen
        Info_TWSE_UD.font = ftGreen
        Info_TWSE_UDR.font = ftGreen
    
    # 美股上色
    if Info_INDU_UDR.value[0] == "-":
        Info_INDU_Price.font = ftGreen
        Info_INDU_UDR.font = ftGreen
    else:
        Info_INDU_Price.font = ftRed
        Info_INDU_UDR.font = ftRed
    if Info_NAS_UDR.value[0] == "-":
        Info_NAS_Price.font = ftGreen
        Info_NAS_UDR.font = ftGreen
    else:
        Info_NAS_Price.font = ftRed
        Info_NAS_UDR.font = ftRed
    if Info_SP500_UDR.value[0] == "-":
        Info_SP500_Price.font = ftGreen
        Info_SP500_UDR.font = ftGreen
    else:
        Info_SP500_Price.font = ftRed
        Info_SP500_UDR.font = ftRed
    if Info_SOX_UDR.value[0] == "-":
        Info_SOX_Price.font = ftGreen
        Info_SOX_UDR.font = ftGreen
    else:
        Info_SOX_Price.font = ftRed
        Info_SOX_UDR.font = ftRed
    
    wb.save(file_name)
    
    
# 存到EXCEL檔中
def Excel(yyyy, mm, dd):
    date = yyyy + "/" + mm + "/" + dd
    
    file_name = "Stock Analysis.xlsx"
    stock_data = [date, str(TWSE_Price), str(TWSE_UD), str(TWSE_UDR) + " %", str(TWSE_Vol) + " 億", str(TWSE_FBS) + " 億", str(INDU_Price), str(INDU_UDR) + " %",
                  str(NAS_Price), str(NAS_UDR) + " %",  str(SP500_Price), str(SP500_UDR) + " %",  str(SOX_Price), str(SOX_UDR) + " %", str(USDEx), str(USDEx_UD), 
                  str(FU), str(DueFu), str(ALL_5_Bull), str(ALL_10_Bull), str(TM_5_Bull), str(TM_10_Bull),
                  str(int(ALL_5_Bull)-int(TM_5_Bull)), str(int(ALL_10_Bull)-int(TM_10_Bull)),
                  str(Buy_Call) + " / " + str(Buy_Put), str(PC_Ratio), str(RIBS_Ratio) + "%"]

    if not os.path.exists(file_name):
        header = ["日期", 
                  "加權指數", "漲跌", "漲跌幅", "成交量", "外資買賣超", 
                  "道瓊工業指數", "漲跌幅", "Nasdaq指數", "漲跌幅", "S&P 500指數", "漲跌幅", "費城半導體指數", "漲跌幅", 
                  "新台幣匯率", "漲跌幅", 
                  "外資期貨未平倉口數", "與結算日相比", "前5大交易人留倉部位(所有)", "前10大交易人留倉部位(所有)", "前5大交易人留倉部位(當月)", "前10大交易人留倉部位(當月)", "前5大交易人未來指標(所有-當月)", "前10大交易人未來指標(所有-當月)",
                  "外資選擇權未平倉金額(買權/賣權)", "P/C Ratio 未平倉量", 
                  "散戶多空比"]
        wb = Workbook()
        ws = wb.active
        ws.title = "台股趨勢"
        ws.append(header)
        ws.append(stock_data)
        wb.save(file_name)
        Excel_color(2)
        
    else:
        date_data_list = []
        wb = load_workbook(file_name)
        ws = wb.get_sheet_by_name("台股趨勢")
        nrow = ws.max_row
        for i in range(2, nrow+1):
            date_data_list.append(ws.cell(row=i, column=1).value)
        # 檢查日期是否存在
        if date in date_data_list:
            choice = input("該日期資料已存在，確認是否覆蓋？(Y/N): ")
            if choice == "Y" or choice =="y":
                change_row = date_data_list.index(date)
                wsw = wb.active
                wsw.title = "台股趨勢"
                for i in range(len(stock_data)):
                    wsw.cell(row=change_row+2, column=i+1).value = stock_data[i]
                wb.save(file_name)
                Excel_color(change_row+2)
        else:
            wsw = wb.active
            wsw.title = "台股趨勢"
            wsw.append(stock_data)
            wb.save(file_name)
            Excel_color(nrow+1)
        
# Menu
def Menu():
    print(" ")
    print("S: 依日期查詢")
    print("E: 離開")
        
def main():
    global INDU_Price, INDU_UDR, NAS_Price, NAS_UDR, SP500_Price, SP500_UDR, SOX_Price, SOX_UDR, USDEx, USDEx_UD
    print("      台股趨勢分析統計資料      ")
    print("-------------------------------")
    
    while True:
        Menu()
        choiceMenu = input("==> ")
        if choiceMenu == "S" or choiceMenu == "s":
            print("日期格式: yyyy/mm/dd")
            yyyy = input("年(西元yyyy): ")
            mm = input("月(mm): ")
            dd = input("日(dd): ")
            if len(mm) == 1:
                mm = "0" + mm
            if len(dd) == 1:
                dd = "0" + dd
            
            # 判斷日期是否有效
            try:
                time.strptime(yyyy + " " + mm + " " + dd, "%Y %m %d")
            except:
                print("日期錯誤\n")
                continue
            
            # 判斷當日有無開盤交易
            YNMarket, stat = Y_N_Market(yyyy, mm, dd)
            if YNMarket == 1:
                print("-------------------------------")
                if stat[0] == "很":
                    print("{}/{}/{} 非交易日".format(yyyy, mm, dd))
                else:
                    print(stat)
                print("-------------------------------\n")
            else:
                try:
                    USASEx(yyyy, mm, dd)
                    USD_NTD(yyyy, mm, dd)
                except:
                    print("當日美股尚未收盤，明日請記得更新美股及新台幣匯率資訊")
                    INDU_Price = 0        # 道瓊工業指數
                    INDU_UDR = 0          # 道瓊工業指數漲跌幅
                    NAS_Price = 0         # Nasdaq指數
                    NAS_UDR = 0           # Nasdaq指數漲跌幅
                    SP500_Price = 0       # S&P 500指數
                    SP500_UDR = 0         # S&P 500指數漲跌幅
                    SOX_Price = 0         # 費城半導體指數
                    SOX_UDR = 0           # 費城半導體指數漲跌幅
                    USDEx = 0             # 美金兌台幣匯率
                    USDEx_UD = 0          # 美金兌台幣匯率漲跌
                    
                TxDueData(yyyy, mm, dd)
                TWSE(yyyy, mm, dd)
                TWFUOC(yyyy, mm, dd)
                TWF_5_10_UOC(yyyy, mm, dd)
                TWMTX(yyyy, mm, dd)
                TWOP(yyyy, mm, dd)
                TWPCR(yyyy, mm, dd)
                Disp(yyyy, mm, dd)
                    
                    
                choiceSave = input("\n是否儲存資料？(Y/N): ")
                if choiceSave == "Y" or choiceSave == "y":
                    Excel(yyyy, mm, dd)
        else:
            break


if __name__ == '__main__':
    main()