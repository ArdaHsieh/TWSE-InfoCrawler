#!/usr/bin/python
# -*- coding: utf-8 -*-
"""
Created on Sat Jul  6 04:36:51 2019

@author: arda
"""

from openpyxl import load_workbook
import pymysql

db = pymysql.connect("127.0.0.1","root", '',"stock" )
cursor = db.cursor()
    
def insert(data): 
    sql = "INSERT INTO trend \
           (Date, \
            TWSE_Price, TWSE_UD, TWSE_UDR, TWSE_Vol, \
            TWSE_FBS, FU, DueFu, Buy_Call, Buy_Put, \
            ALL_5_Bull, ALL_10_Bull, TM_5_Bull, TM_10_Bull, \
            PC_Ratio, RIBS_Ratio, \
            INDU_Price, INDU_UDR, NAS_Price, NAS_UDR, \
            SP500_Price, SP500_UDR, SOX_Price, SOX_UDR, \
            USD2NTDEx, USD2NTDEx_UD) \
           VALUES ('%s', '%s', %s, '%s', %s, '%s', '%s', %s, '%s', %s, \
                   '%s', '%s', %s, '%s', %s, '%s', '%s', %s, '%s', %s, \
                   '%s', '%s', %s, '%s', %s, '%s')" % \
           (data[0][0:4] + '-' + data[0][5:7] + '-' + data[0][8:10], 
            float(data[1]), float(data[2]), float(data[3].split()[0]), float(data[4].split()[0]),
            float(data[5].split()[0]), float(data[16]), float(data[17]), float(data[24].split('/')[0].strip()), float(data[24].split('/')[1].strip()), 
            float(data[18]), float(data[19]), float(data[20]), float(data[21]),
            float(data[25]), float(data[26].split('%')[0]),
            float(data[6]), float(data[7].split()[0]), float(data[8]), float(data[9].split()[0]),
            float(data[10]), float(data[11].split()[0]), float(data[12]), float(data[13].split()[0]),
            float(data[14]), float(data[15]))

    try:
        cursor.execute(sql)
        db.commit()
    except Exception as e:
        db.rollback()
        print(e)
    

file_name = "./result_data/Stock_Analysis.xlsx"

wb = load_workbook(file_name)
ws = wb["台股趨勢"]


for i in range(3, ws.max_row+1):
    row = []
    for j in range(1, 28):
        if ws.cell(row=i, column=j).value == '休市':
            row.append(999999999)
        elif ws.cell(row=i, column=j).value == '結算日':
            row.append(999999999)
        else:
            row.append(ws.cell(row=i, column=j).value)
    insert(row)

db.close()
