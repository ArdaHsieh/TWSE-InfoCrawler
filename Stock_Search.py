#!/usr/bin/python
# -*- coding: utf-8 -*-
"""
Created on Thu Nov 23 2017

@author: I-Ta Hsieh(Arda)
"""


import requests
import json
import os
import datetime
import time
import random
import pymysql
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Color
from operator import itemgetter


day_pre = []               # 前9個交易日
Nth = 0                    # N檔股票    

FO_SE_NBuyNth = []         # 外資買超前30名(上市)
FO_SE_NSellNth = []        # 外資賣超前30名(上市)
FO_SE_NBuy = []            # 外資買超(上市)
FO_SE_NSell = []           # 外資賣超(上市)
FO_SE_ContiNBuy = []       # 外資連續買超(上市)
FO_SE_ContiNSell = []      # 外資連續賣超(上市)
FO_same_IT_SE_NBuy = []    # 外資買超投信相同(上市)
FO_same_IT_SE_NSell = []   # 外資賣超投信相同(上市)
FO_diff_IT_SE_NBuy = []    # 外資買超投信相反(上市)
FO_diff_IT_SE_NSell = []   # 外資賣超投信相反(上市)

IT_SE_NBuyNth = []         # 投信買超前30名(上市)
IT_SE_NSellNth = []        # 投信賣超前30名(上市)
IT_SE_NBuy = []            # 投信買超(上市)
IT_SE_NSell = []           # 投信賣超(上市)
IT_SE_ContiNBuy = []       # 投信連續買超(上市)
IT_SE_ContiNSell = []      # 投信連續賣超(上市)
IT_same_FO_SE_NBuy = []    # 投信買超外資相同(上市)
IT_same_FO_SE_NSell = []   # 投信賣超外資相同(上市)
IT_diff_FO_SE_NBuy = []    # 投信買超外資相反(上市)
IT_diff_FO_SE_NSell = []   # 投信賣超外資相反(上市)

FO_OC_NBuyNth = []         # 外資買超前30名(上櫃)
FO_OC_NSellNth = []        # 外資賣超前30名(上櫃)
FO_OC_NBuy = []            # 外資買超(上櫃)
FO_OC_NSell = []           # 外資賣超(上櫃)
FO_OC_ContiNBuy = []       # 外資連續買超(上櫃)
FO_OC_ContiNSell = []      # 外資連續賣超(上櫃)
FO_same_IT_OC_NBuy = []    # 外資買超投信相同(上櫃)
FO_same_IT_OC_NSell = []   # 外資賣超投信相同(上櫃)
FO_diff_IT_OC_NBuy = []    # 外資買超投信相反(上櫃)
FO_diff_IT_OC_NSell = []   # 外資賣超投信相反(上櫃)

IT_OC_NBuyNth = []         # 投信買超前30名(上櫃)
IT_OC_NSellNth = []        # 投信賣超前30名(上櫃)
IT_OC_NBuy = []            # 投信買超(上櫃)
IT_OC_NSell = []           # 投信賣超(上櫃)
IT_OC_ContiNBuy = []       # 投信連續買超(上櫃)
IT_OC_ContiNSell = []      # 投信連續賣超(上櫃)
IT_same_FO_OC_NBuy = []    # 投信買超外資相同(上櫃)
IT_same_FO_OC_NSell = []   # 投信賣超外資相同(上櫃)
IT_diff_FO_OC_NBuy = []    # 投信買超外資相反(上櫃)
IT_diff_FO_OC_NSell = []   # 投信賣超外資相反(上櫃)


headers = {'user-agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/52.0.2743.116 Safari/537.36'}


def get_proxies():
    url_proxies = "https://free-proxy-list.net/"
    html = requests.get(url_proxies, headers = headers).text.encode('utf-8-sig')
    
    soup = BeautifulSoup(html, 'html.parser')
    tbody = soup.find_all('tbody')
    tr = tbody[0].find_all('tr')
    
    x = random.randint(0, len(tr)-1)
    
    return tr[x].find_all('td')[0].text + ':' + tr[x].find_all('td')[1].text
    
    
def get_url(url):
    html = requests.get(url, headers = headers).text
    
    return html


def post_url(url, payloads):
    html = requests.post(url, headers = headers, data = payloads).text
    
    return html

    
def string_to_nums(text):
    number = ""
    numbers = ["1","2","3","4","5","6","7","8","9","0",".","-"]
    for char in  text:
        if char in numbers:
            number += char
        elif char == "\\" or char == " " or char == "(": 
            break
    
    return float(number)


# 判斷當日有無交易
def Y_N_Market(yyyy, mm, dd):
    url_TWSE = "http://www.twse.com.tw/fund/TWT38U?response=json&date=" + yyyy + mm + dd
    html_TWSE = get_url(url_TWSE)
    TWSE_data1 = json.loads(html_TWSE)
    
    return len(TWSE_data1), TWSE_data1["stat"]


# 判斷前9個交易日
def pre9_day(yyyy, mm, dd):
    global day_pre
    
    f = open('trade_date.txt', 'r')
    Date = list(map(lambda x: x.strip(), f.readlines()))
    n = Date.index(yyyy + '-' + mm + '-' + dd)
    
    for i in range(1, 10):
        day_pre.append(Date[n-i].split('-'))
        

# 10天內首次進榜
def First_in_10(T_Stocks, P_Stocks):
    for stock in T_Stocks:
        if stock[0].strip() in P_Stocks:
            stock.append(0)
        else:
            stock.append(1)
            
    return T_Stocks
            

# 當日買賣超前30名資訊(股號/股名/買賣超股數)
def Net_Buy_Sell_Nth(Tunple_data):
    NBuy = []
    NSell = []
    
    for data in Tunple_data:
        
        if string_to_nums(data[5]) > 0 and len(NBuy) < Nth:
            if (len(data[2].strip()) <= 4 or "KY" in data[2]) and len(data[1].strip()) == 4:
                stock_info = [data[1], data[2], data[5]]
                for stock in TWSE_stocks:
                    if (stock[0].strip() == data[1].strip()):
                        stock_info.append(stock[8])
                        
                        if 'red' in stock[9]:
                            stock_info.append('+' + stock[10])
                        elif 'green' in stock[9]:
                            stock_info.append('-' + stock[10])
                        else:
                            stock_info.append(stock[10])
                        
                        stock_info.append(abs(float(''.join(stock[-2].split(',')))) / float(''.join(stock[2].split(','))))
                        stock_info.append(abs(float(''.join(stock[-1].split(',')))) / float(''.join(stock[2].split(','))))
                        
                        break
                    else:
                        pass
                            
                NBuy.append(stock_info)
        elif string_to_nums(data[5]) <= 0:
            break
    
    for data in Tunple_data:
        if string_to_nums(data[5]) < 0 and len(NSell) < Nth:
            if (len(data[2].strip()) <= 4 or "KY" in data[2]) and len(data[1].strip()) == 4:
                stock_info = [data[1], data[2], data[5]]
                for stock in TWSE_stocks:
                    if (stock[0].strip() == data[1].strip()):
                        stock_info.append(stock[8])
                        
                        if 'red' in stock[9]:
                            stock_info.append('+' + stock[10])
                        elif 'green' in stock[9]:
                            stock_info.append('-' + stock[10])
                        else:
                            stock_info.append(stock[10])
                        
                        stock_info.append(abs(float(''.join(stock[-2].split(',')))) / float(''.join(stock[2].split(','))))
                        stock_info.append(abs(float(''.join(stock[-1].split(',')))) / float(''.join(stock[2].split(','))))
                         
                        break
                    else:
                        pass
                    
                    
                NSell.append(stock_info)

    return NBuy, NSell
    
    
# 買賣超股票
def Net_Buy_Sell_name(Tunple_data):
    NBuy = []
    NSell = []
    
    for data in Tunple_data:
        if string_to_nums(data[5]) > 0:
            NBuy.append(data[2])
        elif string_to_nums(data[5]) < 0:
            NSell.append(data[2])
    
    return NBuy, NSell
 

# 連續買賣超
def Continue(NBuy, NSell, Buy_1, Sell_1, Buy_2, Sell_2) :
    NBuyNth = []
    NSellNth = []
    conti_Buy1 = []
    conti_Buy2 = []
    conti_Sell1 = []
    conti_Sell2 = []
    
    for stock in NBuy:
        if len(NBuyNth) < Nth:
            if len(stock.strip()) <= 4 or ("KY" in stock.strip()):
                NBuyNth.append(stock)
    
    for stock in NSell:
        if len(NSellNth) < Nth:
            if len(stock.strip()) <= 4 or ("KY" in stock.strip()):
                NSellNth.append(stock)
    
    for stock in NBuyNth:
        if stock in Buy_1:
            conti_Buy1.append(stock)
    for stock in conti_Buy1:
        if stock in Buy_2:
            conti_Buy2.append(stock)
            
    for stock in NSellNth:
        if stock in Sell_1:
            conti_Sell1.append(stock)
    for stock in conti_Sell1:
        if stock in Sell_2:
            conti_Sell2.append(stock)
            
    return conti_Buy2, conti_Sell2


# 投信外資反向
def FO_IT_Same_Diff(NBuy1, NSell1, NBuy2, NSell2) :
    Buy_Same = []
    Sell_Same = []
    Buy_Diff = []
    Sell_Diff = []
    NBuyNth = []
    NSellNth = []
    
    for stock in NBuy1:
        if len(NBuyNth) < Nth:
            if len(stock.strip()) <= 4 or ("KY" in stock.strip()):
                NBuyNth.append(stock)
    
    for stock in NSell1:
        if len(NSellNth) < Nth:
            if len(stock.strip()) <= 4 or ("KY" in stock.strip()):
                NSellNth.append(stock)
    
    for stock in NBuyNth:
        if stock in NSell2:
            Buy_Diff.append(stock)
        elif stock in NBuy2:
            Buy_Same.append(stock)
            
    for stock in NSellNth:
        if stock in NSell2:
            Sell_Same.append(stock)
        elif stock in NBuy2:
            Sell_Diff.append(stock)
            
    return Buy_Same, Sell_Same, Buy_Diff, Sell_Diff    
    

# 當日外資買賣超資料(上市)
def FOSE_Info(yyyy, mm, dd):
    url_TWSE = "http://www.twse.com.tw/fund/TWT38U?response=json&date=" + yyyy + mm + dd
    html_TWSE = get_url(url_TWSE)
    TWSE_data1 = json.loads(html_TWSE)
    
    return TWSE_data1['data']


# 當日投信買賣超資料(上市)
def ITSE_Info(yyyy, mm, dd):
    url_TWSE = "http://www.twse.com.tw/fund/TWT44U?response=json&date=" + yyyy + mm + dd
    html_TWSE = get_url(url_TWSE)
    TWSE_data1 = json.loads(html_TWSE)
    
    return TWSE_data1['data']


# 外資上市
def FOSE(yyyy,mm,dd):
    #global day_pre
    global FO_SE_NBuyNth, FO_SE_NSellNth, FO_SE_NBuy, FO_SE_NSell, FO_SE_ContiNBuy, FO_SE_ContiNSell
    
    data = []
    data.append(FOSE_Info(yyyy, mm, dd))

    for i in range(2):
        data.append(FOSE_Info(day_pre[i][0], day_pre[i][1], day_pre[i][2]))
    FO_SE_NBuy, FO_SE_NSell = Net_Buy_Sell_name(data[0])
    Buy_1, Sell_1 = Net_Buy_Sell_name(data[1])
    Buy_2, Sell_2 = Net_Buy_Sell_name(data[2])
    
    FO_SE_ContiNBuy, FO_SE_ContiNSell = Continue(FO_SE_NBuy, FO_SE_NSell, Buy_1, Sell_1, Buy_2, Sell_2)
    FO_SE_NBuyNth, FO_SE_NSellNth = Net_Buy_Sell_Nth(data[0]) 
    
    FO_SE_NBuyNth = First_in_10(FO_SE_NBuyNth, FO_SE_NBuy_Pre_9)
    FO_SE_NSellNth = First_in_10(FO_SE_NSellNth, FO_SE_NSell_Pre_9)
    
    
# 投信上市
def ITSE(yyyy,mm,dd):
    #global day_pre
    global IT_SE_NBuyNth, IT_SE_NSellNth, IT_SE_NBuy, IT_SE_NSell, IT_SE_ContiNBuy, IT_SE_ContiNSell
    
    data = []
    data.append(ITSE_Info(yyyy, mm, dd))
    
    for i in range(2):
        data.append(ITSE_Info(day_pre[i][0], day_pre[i][1],day_pre[i][2]))
    IT_SE_NBuy, IT_SE_NSell = Net_Buy_Sell_name(data[0])
    Buy_1, Sell_1 = Net_Buy_Sell_name(data[1])
    Buy_2, Sell_2 = Net_Buy_Sell_name(data[2])
    
    IT_SE_ContiNBuy, IT_SE_ContiNSell = Continue(IT_SE_NBuy, IT_SE_NSell, Buy_1, Sell_1, Buy_2, Sell_2)
    IT_SE_NBuyNth, IT_SE_NSellNth = Net_Buy_Sell_Nth(data[0])
    
    IT_SE_NBuyNth = First_in_10(IT_SE_NBuyNth, IT_SE_NBuy_Pre_9)
    IT_SE_NSellNth = First_in_10(IT_SE_NSellNth, IT_SE_NSell_Pre_9)
    

# 上市外資投信同步反向
def SE_Same_Diff_way():
    global FO_same_IT_SE_NBuy, FO_same_IT_SE_NSell, FO_diff_IT_SE_NBuy, FO_diff_IT_SE_NSell
    global IT_same_FO_SE_NBuy, IT_same_FO_SE_NSell, IT_diff_FO_SE_NBuy, IT_diff_FO_SE_NSell
    
    FO_same_IT_SE_NBuy, FO_same_IT_SE_NSell, FO_diff_IT_SE_NBuy, FO_diff_IT_SE_NSell = FO_IT_Same_Diff(FO_SE_NBuy, FO_SE_NSell, IT_SE_NBuy, IT_SE_NSell)
    IT_same_FO_SE_NBuy, IT_same_FO_SE_NSell, IT_diff_FO_SE_NBuy, IT_diff_FO_SE_NSell = FO_IT_Same_Diff(IT_SE_NBuy, IT_SE_NSell, FO_SE_NBuy, FO_SE_NSell)
    

# 所有上市個股資料
def TWSE_All_Stocks(yyyy, mm, dd):
    global TWSE_stocks
    url_TWSE = "http://www.tse.com.tw/exchangeReport/MI_INDEX?response=text&date=https://www.twse.com.tw/exchangeReport/MI_INDEX?response=json&date=" + yyyy + mm + dd + "&type=ALL"
    html_TWSE = get_url(url_TWSE)
    TWSE_data = json.loads(html_TWSE)['data9']
    
    TWSE_stocks = list(filter(lambda x: x[0][0] != '0' and x[0][-1].isdigit() and len(x[0]) == 4, TWSE_data))
    for s in TWSE_stocks:
        s.extend(['0','0'])

    FOSE_Data = FOSE_Info(yyyy, mm, dd)
    ITSE_Data = ITSE_Info(yyyy, mm, dd)
    FOSE_Data = sorted(list(filter(lambda x: len(x[1].strip()) == 4, FOSE_Data)), key=itemgetter(1))
    ITSE_Data = sorted(list(filter(lambda x: len(x[1].strip()) == 4, ITSE_Data)), key=itemgetter(1))
    
    for stock in TWSE_stocks:
        for s in FOSE_Data:
            if stock[0].strip() == s[1].strip():
                stock[-2] = s[5]
                break
        for s in ITSE_Data:
            if stock[0].strip() == s[1].strip():
                stock[-1] = s[5]
                break


# 上櫃當日買賣超前30名資訊(股號/股名/買賣超股數)
def OC_Net_Buy_Sell_Nth(Tunple_data):
    NStock = []
    
    for data in Tunple_data:
        if len(NStock) < Nth:
            for stock in TWOC_stocks:
                if data[0] == stock[0]:
                    data.append(stock[2])
                    data.append(stock[3])
                    data.append(abs(float(''.join(stock[-2].split(','))))*1000 / float(''.join(stock[7].split(','))))
                    data.append(abs(float(''.join(stock[-1].split(','))))*1000 / float(''.join(stock[7].split(','))))
                    NStock.append(data)
                    break
                else:
                    pass
        else:
            break

    return NStock


# 當日外資買賣超資料(上櫃)
def FOOC_Info(yyyy, mm, dd):
    TWOC_Buy, TWOC_Sell = [], []
    url_TWOC_Buy = "https://www.tpex.org.tw/web/stock/3insti/qfii_trading/forgtr_result.php?l=zh-tw&t=D&type=buy&d=" + str(int(yyyy)-1911) + '/' + mm + '/' + dd
    url_TWOC_Sell = "https://www.tpex.org.tw/web/stock/3insti/qfii_trading/forgtr_result.php?l=zh-tw&t=D&type=sell&d=" + str(int(yyyy)-1911) + '/' + mm + '/' + dd
    html_TWOC_Buy = get_url(url_TWOC_Buy)
    html_TWOC_Sell = get_url(url_TWOC_Sell)
    
    TWOC_data = list(filter(lambda x: len(x[1]) == 4, json.loads(html_TWOC_Buy)['aaData']))
    
    for datalist in TWOC_data:
        oc_data = [datalist[1], datalist[2].replace(" ", ""), datalist[5]]
        TWOC_Buy.append(oc_data)
        
    
    TWOC_data = list(filter(lambda x: len(x[1]) == 4, json.loads(html_TWOC_Sell)['aaData']))
    
    for datalist in TWOC_data:
        oc_data = [datalist[1], datalist[2].replace(" ", ""), datalist[5]]
        TWOC_Sell.append(oc_data)
    
    return [TWOC_Buy, TWOC_Sell]


# 當日投信買賣超資料(上櫃)
def ITOC_Info(yyyy, mm, dd):
    TWOC_Buy, TWOC_Sell = [], []
    url_TWOC_Buy = "https://www.tpex.org.tw/web/stock/3insti/sitc_trading/sitctr_result.php?l=zh-tw&t=D&type=buy&d=" + str(int(yyyy)-1911) + '/' + mm + '/' + dd
    url_TWOC_Sell = "https://www.tpex.org.tw/web/stock/3insti/sitc_trading/sitctr_result.php?l=zh-tw&t=D&type=sell&d=" + str(int(yyyy)-1911) + '/' + mm + '/' + dd
    html_TWOC_Buy = get_url(url_TWOC_Buy)
    html_TWOC_Sell = get_url(url_TWOC_Sell)
    
    TWOC_data = list(filter(lambda x: len(x[1]) == 4, json.loads(html_TWOC_Buy)['aaData']))
    
    for datalist in TWOC_data:
        oc_data = [datalist[1], datalist[2].replace(" ", ""), datalist[5]]
        TWOC_Buy.append(oc_data)
        
    
    TWOC_data = list(filter(lambda x: len(x[1]) == 4, json.loads(html_TWOC_Sell)['aaData']))
    
    for datalist in TWOC_data:
        oc_data = [datalist[1], datalist[2].replace(" ", ""), datalist[5]]
        TWOC_Sell.append(oc_data)
    
    return [TWOC_Buy, TWOC_Sell]


# 外資上櫃
def FOOC(yyyy,mm,dd):
    #global day_pre
    global FO_OC_NBuyNth, FO_OC_NSellNth, FO_OC_NBuy, FO_OC_NSell, FO_OC_ContiNBuy, FO_OC_ContiNSell
    
    data = []
    data.append(FOOC_Info(yyyy, mm, dd))

    for i in range(2):
        data.append(FOOC_Info(day_pre[i][0], day_pre[i][1], day_pre[i][2]))
    FO_OC_NBuy, FO_OC_NSell = [stock[1] for stock in data[0][0]], [stock[1] for stock in data[0][1]]
    Buy_1, Sell_1 = [stock[1] for stock in data[1][0]], [stock[1] for stock in data[1][1]]
    Buy_2, Sell_2 = [stock[1] for stock in data[2][0]], [stock[1] for stock in data[2][1]]
    
    FO_OC_ContiNBuy, FO_OC_ContiNSell = Continue(FO_OC_NBuy, FO_OC_NSell, Buy_1, Sell_1, Buy_2, Sell_2)
    FO_OC_NBuyNth, FO_OC_NSellNth = OC_Net_Buy_Sell_Nth(data[0][0]), OC_Net_Buy_Sell_Nth(data[0][1]) 
    
    FO_OC_NBuyNth = First_in_10(FO_OC_NBuyNth, FO_OC_NBuy_Pre_9)
    FO_OC_NSellNth = First_in_10(FO_OC_NSellNth, FO_OC_NSell_Pre_9)
    
    
# 投信上櫃
def ITOC(yyyy,mm,dd):
    #global day_pre
    global IT_OC_NBuyNth, IT_OC_NSellNth, IT_OC_NBuy, IT_OC_NSell, IT_OC_ContiNBuy, IT_OC_ContiNSell
    
    data = []
    data.append(ITOC_Info(yyyy, mm, dd))
    
    for i in range(2):
        data.append(ITOC_Info(day_pre[i][0], day_pre[i][1], day_pre[i][2]))
    IT_OC_NBuy, IT_OC_NSell = [stock[1] for stock in data[0][0]], [stock[1] for stock in data[0][1]]
    Buy_1, Sell_1 = [stock[1] for stock in data[1][0]], [stock[1] for stock in data[1][1]]
    Buy_2, Sell_2 = [stock[1] for stock in data[2][0]], [stock[1] for stock in data[2][1]]
    
    IT_OC_ContiNBuy, IT_OC_ContiNSell = Continue(IT_OC_NBuy, IT_OC_NSell, Buy_1, Sell_1, Buy_2, Sell_2)
    IT_OC_NBuyNth, IT_OC_NSellNth = OC_Net_Buy_Sell_Nth(data[0][0]), OC_Net_Buy_Sell_Nth(data[0][1])  
    
    IT_OC_NBuyNth = First_in_10(IT_OC_NBuyNth, IT_OC_NBuy_Pre_9)
    IT_OC_NSellNth = First_in_10(IT_OC_NSellNth, IT_OC_NSell_Pre_9)
    
    
# 上櫃外資投信同步反向
def OC_Same_Diff_way():
    global FO_same_IT_OC_NBuy, FO_same_IT_OC_NSell, FO_diff_IT_OC_NBuy, FO_diff_IT_OC_NSell
    global IT_same_FO_OC_NBuy, IT_same_FO_OC_NSell, IT_diff_FO_OC_NBuy, IT_diff_FO_OC_NSell
    
    FO_same_IT_OC_NBuy, FO_same_IT_OC_NSell, FO_diff_IT_OC_NBuy, FO_diff_IT_OC_NSell = FO_IT_Same_Diff(FO_OC_NBuy, FO_OC_NSell, IT_OC_NBuy, IT_OC_NSell)
    IT_same_FO_OC_NBuy, IT_same_FO_OC_NSell, IT_diff_FO_OC_NBuy, IT_diff_FO_OC_NSell = FO_IT_Same_Diff(IT_OC_NBuy, IT_OC_NSell, FO_OC_NBuy, FO_OC_NSell)
    

# 所有上櫃個股資料
def TWOC_All_Stocks(yyyy, mm, dd):
    global TWOC_stocks
    url_TWOC = "https://www.tpex.org.tw/web/stock/aftertrading/otc_quotes_no1430/stk_wn1430_result.php?l=zh-tw&d=" + str(int(yyyy)-1911) + '/' + mm + '/' + dd + "&se=AL"
    html_TWOC = get_url(url_TWOC)
    TWOC_data = json.loads(html_TWOC)['aaData']
    
    TWOC_stocks = list(filter(lambda x: len(x[0]) == 4, TWOC_data))
    
    for s in TWOC_stocks:
        s.extend(['0','0'])

    FOOC_Buy_Sell = FOOC_Info(yyyy,mm,dd)
    ITOC_Buy_Sell = ITOC_Info(yyyy,mm,dd)
    
    for stock in TWOC_stocks:
        for s in FOOC_Buy_Sell[0]:
            if stock[0].strip() == s[0].strip():
                stock[-2] = s[2]
                break
        for s in FOOC_Buy_Sell[1]:
            if stock[0].strip() == s[0].strip():
                stock[-2] = s[2]
                break
        for s in ITOC_Buy_Sell[0]:
            if stock[0].strip() == s[0].strip():
                stock[-1] = s[2]
                break
        for s in ITOC_Buy_Sell[1]:
            if stock[0].strip() == s[0].strip():
                stock[-1] = s[2]
                break
    
    
    
# 存到EXCEL檔中
def Excel(filename, sheet_name, NetNth, NetNth_Same, NetNth_Diff, Conti):
    wb = load_workbook(filename)
    #ws = wb.get_sheet_by_name(sheet_name)
    ws = wb[sheet_name]

    for i in range(len(NetNth)):
        ws.cell(row=i+2, column=1).value = NetNth[i][0]
        ws.cell(row=i+2, column=2).value = NetNth[i][1]
        ws.cell(row=i+2, column=3).value = NetNth[i][2]
        ws.cell(row=i+2, column=4).value = NetNth[i][3]
        ws.cell(row=i+2, column=5).value = NetNth[i][4]
        ws.cell(row=i+2, column=6).value = round(NetNth[i][5], 4) * 100
        ws.cell(row=i+2, column=7).value = round(NetNth[i][6], 4) * 100
       
    red_fill = PatternFill(patternType='solid', fgColor=Color('FF8888'))
    blue_fill = PatternFill(patternType='solid', fgColor=Color('33CCFF'))
    green_fill = PatternFill(patternType='solid', fgColor=Color('66FF66'))
    yellow_fill = PatternFill(patternType='solid', fgColor=Color('FFFF77'))
    
    for stock in Conti:
        for i in range(len(NetNth)):
            if NetNth[i][1] == stock:
                ws.cell(row=i+2, column=1).fill = blue_fill
                #break
    for i in range(len(NetNth)):
        if NetNth[i][7] == 1:
            ws.cell(row=i+2, column=1).fill = yellow_fill
            #break
    for stock in NetNth_Same:
        for i in range(len(NetNth)):
            if NetNth[i][1] == stock:
                ws.cell(row=i+2, column=2).fill = red_fill
                #break
    for stock in NetNth_Diff:
        for i in range(len(NetNth)):
            if NetNth[i][1] == stock:
                ws.cell(row=i+2, column=2).fill = green_fill
    
    ws.cell(row=1, column=1).value = "證券代號"
    ws.cell(row=1, column=2).value = "證券名稱"
    ws.cell(row=1, column=3).value = "張數"
    ws.cell(row=1, column=4).value = "收盤價"
    ws.cell(row=1, column=5).value = "漲跌"
    ws.cell(row=1, column=6).value = "外資交易佔比"
    ws.cell(row=1, column=7).value = "投信交易佔比"
    ws.cell(row=1, column=9).value = "連續3天"
    ws.cell(row=1, column=10).value = "10天內首次"
    ws.cell(row=1, column=11).value = "同向"
    ws.cell(row=1, column=12).value = "反向"
    ws.cell(row=1, column=9).fill = blue_fill
    ws.cell(row=1, column=10).fill = yellow_fill
    ws.cell(row=1, column=11).fill = red_fill
    ws.cell(row=1, column=12).fill = green_fill    
    
    wb.save(filename)


# 存到DB中
def DB(Date, StockType, ActType, NetNth, NetNth_Same, NetNth_Diff, Conti):
    db = pymysql.connect("127.0.0.1","root", '',"stock" )
    cursor = db.cursor()
        
    date = Date[0] + "-" + Date[1] + "-" + Date[2]
    
    for i in range(len(NetNth)):
        NetNth[i].extend([0,0,0])
              
    for stock in NetNth_Same:
        for i in range(len(NetNth)):
            if NetNth[i][1] == stock:
                NetNth[i][8] = 1
                
    for stock in NetNth_Diff:
        for i in range(len(NetNth)):
            if NetNth[i][1] == stock:
               NetNth[i][8] = -1

    for stock in Conti:
        for i in range(len(NetNth)):
            if NetNth[i][1] == stock:
                NetNth[i][9] = 1
     
    for i in range(len(NetNth)):              
        sql = "INSERT INTO chip_stocks_" + StockType + " \
              (Date, \
               ID, Name, Vol, Price, UD, FO_Vol_Rate, IT_Vol_Rate, \
               Actype, FO_IT_SD, Conti2, Fst_in_10) \
               VALUES ('%s', '%s', %s, '%s', %s, '%s', '%s', %s, '%s', '%s', '%s', '%s')" % \
              (date,
               str(NetNth[i][0]), "'" + str(NetNth[i][1]) + "'", str(NetNth[i][2]), 
               float(''.join(NetNth[i][3].split(','))), 
               str(NetNth[i][4]), 
               float(round(NetNth[i][5]*100, 2)),
               float(round(NetNth[i][6]*100, 2)),
               ActType, int(NetNth[i][8]), int(NetNth[i][9]), int(NetNth[i][7]))

        try:
            cursor.execute(sql)
            db.commit()
        except Exception as e:
            db.rollback()
            print(e)
    
    db.close()


# 從DB找出前9日資料
def pre9_DB():
    global FO_SE_NBuy_Pre_9, FO_SE_NSell_Pre_9
    global IT_SE_NBuy_Pre_9, IT_SE_NSell_Pre_9 
    global FO_OC_NBuy_Pre_9, FO_OC_NSell_Pre_9 
    global IT_OC_NBuy_Pre_9, IT_OC_NSell_Pre_9 
    
    db = pymysql.connect("127.0.0.1","root", '',"stock" )
    cursor = db.cursor()
    
    stop = day_pre[0][0] + '-' + day_pre[0][1] + '-' + day_pre[0][2]
    start = day_pre[-1][0] + '-' + day_pre[-1][1] + '-' + day_pre[-1][2]
    
    # FO_SE_BUY
    sql = "SELECT distinct ID FROM stock.chip_stocks_se \
           WHERE Actype = '%s' AND \
           Date(str_to_date(date,'%%Y-%%m-%%d')) between '%s' AND '%s'" % \
           ('fo_nbuy', start, stop)

    try:
        cursor.execute(sql)
        FO_SE_NBuy_Pre_9 = [item[0].strip() for item in cursor.fetchall()]
    except Exception as e:
        db.rollback()
        print(e)
    
    # FO_SE_SELL
    sql = "SELECT distinct ID FROM stock.chip_stocks_se \
           WHERE Actype = '%s' AND \
           Date(str_to_date(date,'%%Y-%%m-%%d')) between '%s' AND '%s'" % \
           ('fo_nsell', start, stop)

    try:
        cursor.execute(sql)
        FO_SE_NSell_Pre_9 = [item[0].strip() for item in cursor.fetchall()]
    except Exception as e:
        db.rollback()
        print(e)
    
    # IT_SE_BUY
    sql = "SELECT distinct ID FROM stock.chip_stocks_se \
           WHERE Actype = '%s' AND \
           Date(str_to_date(date,'%%Y-%%m-%%d')) between '%s' AND '%s'" % \
           ('it_nbuy', start, stop)

    try:
        cursor.execute(sql)
        IT_SE_NBuy_Pre_9 = [item[0].strip() for item in cursor.fetchall()]
    except Exception as e:
        db.rollback()
        print(e)
    
    # IT_SE_SELL
    sql = "SELECT distinct ID FROM stock.chip_stocks_se \
           WHERE Actype = '%s' AND \
           Date(str_to_date(date,'%%Y-%%m-%%d')) between '%s' AND '%s'" % \
           ('it_nsell', start, stop)

    try:
        cursor.execute(sql)
        IT_SE_NSell_Pre_9 = [item[0].strip() for item in cursor.fetchall()]
    except Exception as e:
        db.rollback()
        print(e)
    
    # FO_OC_BUY
    sql = "SELECT distinct ID FROM stock.chip_stocks_oc \
           WHERE Actype = '%s' AND \
           Date(str_to_date(date,'%%Y-%%m-%%d')) between '%s' AND '%s'" % \
           ('fo_nbuy', start, stop)

    try:
        cursor.execute(sql)
        FO_OC_NBuy_Pre_9 = [item[0].strip() for item in cursor.fetchall()]
    except Exception as e:
        db.rollback()
        print(e)
    
    # FO_OC_SELL
    sql = "SELECT distinct ID FROM stock.chip_stocks_oc \
           WHERE Actype = '%s' AND \
           Date(str_to_date(date,'%%Y-%%m-%%d')) between '%s' AND '%s'" % \
           ('fo_nsell', start, stop)

    try:
        cursor.execute(sql)
        FO_OC_NSell_Pre_9 = [item[0].strip() for item in cursor.fetchall()]
    except Exception as e:
        db.rollback()
        print(e)
    
    # IT_OC_BUY
    sql = "SELECT distinct ID FROM stock.chip_stocks_oc \
           WHERE Actype = '%s' AND \
           Date(str_to_date(date,'%%Y-%%m-%%d')) between '%s' AND '%s'" % \
           ('it_nbuy', start, stop)

    try:
        cursor.execute(sql)
        IT_OC_NBuy_Pre_9 = [item[0].strip() for item in cursor.fetchall()]
    except Exception as e:
        db.rollback()
        print(e)
    
    # IT_OC_SELL
    sql = "SELECT distinct ID FROM stock.chip_stocks_oc \
           WHERE Actype = '%s' AND \
           Date(str_to_date(date,'%%Y-%%m-%%d')) between '%s' AND '%s'" % \
           ('it_nsell', start, stop)

    try:
        cursor.execute(sql)
        IT_OC_NSell_Pre_9 = [item[0].strip() for item in cursor.fetchall()]
    except Exception as e:
        db.rollback()
        print(e)

    db.close()
    
        
# Menu
def Menu():
    print("S: 依日期查詢")
    print("E: 離開")    
    

def main():
    global Nth
    
    print("          台股籌碼選股          ")
    print("--------------------------------")
    
    while True:
        Menu()
        choiceMenu = input("==> ")
        if choiceMenu == "S" or choiceMenu == "s":
            print("日期格式: yyyy/mm/dd")
            yyyy = input("年(西元yyyy): ")
            mm = input("月(mm): ")
            dd = input("日(dd): ")
            if len(mm) == 1:
                mm = "0" + mm
            if len(dd) == 1:
                dd = "0" + dd
            
            # 判斷日期是否有效
            try:
                time.strptime(yyyy + " " + mm + " " + dd, "%Y %m %d")
            except:
                print("日期錯誤\n")
                continue
            
            # 判斷當日有無開盤交易
            YNMarket, stat = Y_N_Market(yyyy, mm, dd)
            if YNMarket == 1:
                print("-------------------------------")
                if stat[0] == "很":
                    print("{}/{}/{} 非交易日".format(yyyy, mm, dd))
                else:
                    print(stat)
                print("-------------------------------\n")
            # 判斷剛日期檔案是否存在 
            else:
                if not os.path.isdir("./result_data/"):
                    os.mkdir("./result_data/")
                filename = "./result_data/Stock_condidates_" + yyyy + mm + dd + ".xlsx"
                if os.path.exists(filename):
                    print("該日期資料已存在")
                else:
                    TWSE_All_Stocks(yyyy, mm, dd)
                    TWOC_All_Stocks(yyyy, mm, dd)
                    
                    wb = Workbook()
                    wb.active.title = "上市外資買超"
                    wb.create_sheet(title = "上市外資賣超")
                    wb.create_sheet(title = "上市投信買超")
                    wb.create_sheet(title = "上市投信賣超")
                    wb.create_sheet(title = "上櫃外資買超")
                    wb.create_sheet(title = "上櫃外資賣超")
                    wb.create_sheet(title = "上櫃投信買超")
                    wb.create_sheet(title = "上櫃投信賣超")
                    wb.save(filename)
                     
                    Nth = int(input("預記錄之買賣超股數: "))
                    pre9_day(yyyy,mm,dd)
                    pre9_DB()
                    time.sleep(5)
                    FOSE(yyyy,mm,dd)
                    FOOC(yyyy,mm,dd)
                    time.sleep(10)
                    ITSE(yyyy,mm,dd)
                    ITOC(yyyy,mm,dd)
                    time.sleep(5)
                    SE_Same_Diff_way()
                    OC_Same_Diff_way()
                    Excel(filename, "上市外資買超", FO_SE_NBuyNth, FO_same_IT_SE_NBuy, FO_diff_IT_SE_NBuy, FO_SE_ContiNBuy)
                    Excel(filename, "上市外資賣超", FO_SE_NSellNth, FO_same_IT_SE_NSell, FO_diff_IT_SE_NSell, FO_SE_ContiNSell)
                    Excel(filename, "上市投信買超", IT_SE_NBuyNth, IT_same_FO_SE_NBuy, IT_diff_FO_SE_NBuy, IT_SE_ContiNBuy)
                    Excel(filename, "上市投信賣超", IT_SE_NSellNth, IT_same_FO_SE_NSell, IT_diff_FO_SE_NSell, IT_SE_ContiNSell)
                    Excel(filename, "上櫃外資買超", FO_OC_NBuyNth, FO_same_IT_OC_NBuy, FO_diff_IT_OC_NBuy, FO_OC_ContiNBuy)
                    Excel(filename, "上櫃外資賣超", FO_OC_NSellNth, FO_same_IT_OC_NSell, FO_diff_IT_OC_NSell, FO_OC_ContiNSell)
                    Excel(filename, "上櫃投信買超", IT_OC_NBuyNth, IT_same_FO_OC_NBuy, IT_diff_FO_OC_NBuy, IT_OC_ContiNBuy)
                    Excel(filename, "上櫃投信賣超", IT_OC_NSellNth, IT_same_FO_OC_NSell, IT_diff_FO_OC_NSell, IT_OC_ContiNSell)
                    
                    DB([yyyy,mm,dd], "se", "fo_nbuy", FO_SE_NBuyNth, FO_same_IT_SE_NBuy, FO_diff_IT_SE_NBuy, FO_SE_ContiNBuy)
                    DB([yyyy,mm,dd], "se", "fo_nsell", FO_SE_NSellNth, FO_same_IT_SE_NSell, FO_diff_IT_SE_NSell, FO_SE_ContiNSell)
                    DB([yyyy,mm,dd], "se", "it_nbuy", IT_SE_NBuyNth, IT_same_FO_SE_NBuy, IT_diff_FO_SE_NBuy, IT_SE_ContiNBuy)
                    DB([yyyy,mm,dd], "se", "it_nsell", IT_SE_NSellNth, IT_same_FO_SE_NSell, IT_diff_FO_SE_NSell, IT_SE_ContiNSell)
                    DB([yyyy,mm,dd], "oc", "fo_nbuy", FO_OC_NBuyNth, FO_same_IT_OC_NBuy, FO_diff_IT_OC_NBuy, FO_OC_ContiNBuy)
                    DB([yyyy,mm,dd], "oc", "fo_nsell", FO_OC_NSellNth, FO_same_IT_OC_NSell, FO_diff_IT_OC_NSell, FO_OC_ContiNSell)
                    DB([yyyy,mm,dd], "oc", "it_nbuy", IT_OC_NBuyNth, IT_same_FO_OC_NBuy, IT_diff_FO_OC_NBuy, IT_OC_ContiNBuy)
                    DB([yyyy,mm,dd], "oc", "it_nsell", IT_OC_NSellNth, IT_same_FO_OC_NSell, IT_diff_FO_OC_NSell, IT_OC_ContiNSell)

        else:
            break
                     
   
if __name__ == '__main__':
    main()