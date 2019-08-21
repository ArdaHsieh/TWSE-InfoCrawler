#!/usr/bin/python
# -*- coding: utf-8 -*-
"""
Created on Sat Jul  6 04:36:51 2019

@author: arda
"""

from openpyxl import load_workbook
import pymysql

db = pymysql.connect("127.0.0.1","root", '',"stock" )
cursor = db.cursor()
    
def insert(StockType, ActType, date, data): 
    sql = "INSERT INTO chip_stocks_" + StockType + " \
              (Date, \
               ID, Name, Vol, Price, UD, \
               Actype, FO_IT_SD, Conti2, Fst_in_10) \
               VALUES ('%s', '%s', %s, '%s', %s, '%s', '%s', %s, '%s', '%s')" % \
              (date,
              str(data[0]).strip(), "'" + str(data[1]).strip() + "'", str(data[2]).strip(), float(''.join(data[3].split(','))), str(data[4]).strip(),
               ActType, int(data[5]), int(data[6]), 0)


    try:
        cursor.execute(sql)
        db.commit()
    except Exception as e:
        db.rollback()
        print(e)
    

file_name = "./result_data/Stock_condidates_20190719.xlsx"

wb = load_workbook(file_name)
ws = wb["上市外資買超"]


#for i in range(3, ws.max_row+1):
for i in range(3, 6):
    row = []
    for j in range(1, 6):
        #print(str(type(ws.cell(row=i, column=j).value)) + ':' + str(ws.cell(row=i, column=j).value))
        row.append(ws.cell(row=i, column=j).value)
    row.append(-1)
    row.append(0)
    print(type(row[1]))
    print(row[1])
    #print('\n')

    insert('se', 'fo_nbuy', '2019-7-19', row)

db.close()
